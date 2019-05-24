import os
import csv
from openpyxl import load_workbook, Workbook


def _convert_csv_to_xlsx(list_loc):
    # TODO: Test function with actual CSV as it would be provided
    try:
        csv_file = csv.reader(
            open(list_loc, "r"),
            quotechar='"',
            delimiter=",", quoting=csv.QUOTE_ALL,
            skipinitialspace=True)

    except FileNotFoundError:
        print("Is the csv in the 'input' folder? Have you spelled it correctly?")
        return

    else:
        wb = Workbook(write_only=True)
        ws = wb.active
        for row in csv_file:
            ws.append(row)
        return ws


def sanitize_upload(f):
    # If an xls, complain

    if f.endswith(".xls"):
        print('Need the file to be in .xlsx!')
        return  # TODO provide feedback to the uploader

    # If a csv, convert to xlsx and return the worksheet
    elif f.endswith(".csv"):
        ws = _convert_csv_to_xlsx(f)
        if not ws:
            return
        print("Converted from .csv to .xlsx")

    # If an xlsx load the workbook and select the active sheet
    elif f.endswith(".xlsx"):
        ws = load_workbook(f).active
        print("Using an xlsx. You didn't need to convert this yourself. You could have passed in the raw '.csv'")

    # Passed a file that has no business being here
    else:
        _, ext = os.path.splitext(f)
        print(
            f'What are you doing giving this script a {ext}??!! Do you know what this script even does?! Have you read the documentation!  I spent a whole morning writing it!')
        return  # TODO WHAT DO WE DO HERE?
    return ws


def get_sign_up_items(student_list):
    """Loops through the rows in 'student_list' and creates a convenient list for use in populating student model
  Useful Columns are:

  C firstname
  D lastname
  E preferredname
  
  Parameters
  ----------
  student_list : str
    The file location of the spreadsheet, (File type object)
  enrol : bool, optional
    A flag that informs the generation of a list of sign up items, or just emails for enrolment  (default is False)
  print_items: bool, optional
    A flag that controls wether the items are output to the console, so they can be cut from the output and used directly; Used if script is run within the Repl.it (default is False)

  Returns
  -------
  list
    A list of useful items to populate a student model
  
  """

    ws = sanitize_upload(student_list)

    data = []

    for row in range(2, ws.max_row + 1):
        bio_list = []
        for column in "ACDEJ":  # firstname, lastname, preferred,
            cell_name = "{}{}".format(column, row)
            bio_list.append(ws[cell_name].value)

        data.append(bio_list)
    return data


def get_gradebook_items(gradebook):
    assignment_list = []
    student_fyid_list = []
    ws = sanitize_upload(gradebook)
    max_col = ws.max_column
    max_row = ws.max_row
    # The assignments we'd like to assign as part of the ISP
    okList = ['Quiz:', 'Assignment:', 'Interactive Content:']

    for i in range(4, max_col):
        assignment = ws.cell(row=1, column=i).value
        if any(s in assignment for s in okList):
            assignment_list.append(assignment)

    for i in range(2, max_row):
        student_fyid = ws.cell(row=i, column=3).value.split("@")[0]
        student_fyid_list.append(student_fyid.lower())

    return assignment_list, student_fyid_list


def _pan(title):
    if title[0:4] == 'Quiz':
        p_title = title. \
            replace('Quiz:', ''). \
            replace('(Real)', ''). \
            replace('Moodle', ''). \
            replace('Quiz', ''). \
            replace(' :', '')
    elif title[0:10] == "Assignment":
        p_title = title. \
            replace('Assignment:', ''). \
            replace('(Real)', ''). \
            replace('Moodle', ''). \
            replace('Assignment', ''). \
            replace('Portfolio:', ''). \
            replace(' :', '')
    else:
        p_title = title. \
            replace('Interactive Content:', ''). \
            replace('(Real)', '')

    return p_title


def get_student_items(gradebook):
    student_data_list = []
    student_grade_dict = {}
    fyid_lookup = {}
    ws = sanitize_upload(gradebook)
    okList = ['Quiz:', 'Assignment:', 'Interactive Content:']
    max_row = ws.max_row
    max_col = ws.max_column
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=3, max_row=max_row, values_only=True)):

        fyid = row[2].split("@")[0]
        surname = row[1].upper()
        names = row[0].split('(')
        if len(names) > 1:
            given = names[1].replace(')', '')
            pref = f' ({names[0].strip()})'
        else:
            given = names[0]
            pref = ""
        student_grade_dict[i + 2] = fyid
        student_data_list.append((fyid, surname, given, pref))

    grade_dict = {}

    # Slide down the row of names
    for i in range(2, max_row + 1):
        collector = {}
        # For that row, starting at column four iterate through each cell
        for j, cell in enumerate(ws.iter_cols(min_row=i,
                                              max_row=i,
                                              min_col=4,
                                              max_col=max_col - 2,
                                              values_only=True)):
            assignment = ws.cell(row=1, column=j + 4).value

            if any(s in assignment for s in okList):
                if cell[0] not in ['-', None]:
                    collector[_pan(assignment)] = cell[0]
        grade_dict[i] = collector

    for k, v in grade_dict.items():
        fyid_lookup[student_grade_dict[k]] = v

    return student_data_list, fyid_lookup
